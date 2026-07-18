import json
import os
import csv


def read_spreadsheet(args_json):
    args = json.loads(args_json)
    path = args["input_path"]
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = [row for row in reader]
        return json.dumps(rows[:100])
    elif ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_name = args.get("sheet", wb.sheetnames[0])
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            return json.dumps(rows[:100])
        except ImportError:
            return "ERROR: openpyxl not available"
    return f"ERROR: Unsupported format: {ext}"


def create_spreadsheet(args_json):
    args = json.loads(args_json)
    data = json.loads(args.get("data", '{"sheets":[]}'))
    try:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_data in data.get("sheets", []):
            ws = wb.create_sheet(sheet_data.get("name", "Sheet1"))
            headers = sheet_data.get("headers", [])
            if headers:
                ws.append(headers)
            for row in sheet_data.get("rows", []):
                ws.append(row)
        wb.save(args["output_path"])
        return f"Created spreadsheet → {args['output_path']}"
    except ImportError:
        return "ERROR: openpyxl not available"


def csv_to_xlsx(args_json):
    args = json.loads(args_json)
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        with open(args["input_path"], "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        wb.save(args["output_path"])
        return f"Converted CSV → XLSX: {args['output_path']}"
    except ImportError:
        return "ERROR: openpyxl not available"


def spreadsheet_to_pdf(args_json):
    args = json.loads(args_json)
    result = read_spreadsheet(args_json)
    try:
        data = json.loads(result)
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4, landscape
        c = canvas.Canvas(args["output_path"], pagesize=landscape(A4))
        w, h = landscape(A4)
        y = h - 30
        c.setFont("Helvetica", 8)
        for row in data[:50]:
            line = " | ".join([str(cell)[:20] for cell in row])
            if y < 30:
                c.showPage()
                y = h - 30
            c.drawString(20, y, line[:120])
            y -= 12
        c.save()
        return f"Converted spreadsheet → PDF: {args['output_path']}"
    except Exception as e:
        return f"ERROR: {e}"
